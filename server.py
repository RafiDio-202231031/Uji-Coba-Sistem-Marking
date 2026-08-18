# =====================================================================
#  server.py - PYTHON: server, API, keamanan, dan export Excel
#  Jalankan: python server.py  ->  http://localhost:5000
# =====================================================================
import os, re, io, hashlib, sqlite3
from datetime import datetime
from functools import wraps
from flask import Flask, request, jsonify, session, send_from_directory, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.pagebreak import Break
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

BASE = os.path.dirname(os.path.abspath(__file__))
DB   = os.path.join(BASE, "packing_java_volume_art.db")
WARNA_OPSI = ["Natural Waterbase", "Black Burnt", "Black", "Dirty Brown",
              "Rustic", "White Bleached", "Bleached"]

app = Flask(__name__)
app.secret_key = "jva-marking-anti-double-2026"

# ---------------- DATABASE & KEAMANAN ----------------
def hash_password(password, salt=None):
    salt = salt or os.urandom(16)
    h = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 100_000)
    return salt.hex(), h.hex()

def verifikasi_password(password, salt_hex, hash_hex):
    h = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), bytes.fromhex(salt_hex), 100_000)
    return h.hex() == hash_hex

def get_db():
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS packing (
        id INTEGER PRIMARY KEY AUTOINCREMENT, packing_no TEXT NOT NULL,
        shipment_no TEXT NOT NULL, code TEXT, description TEXT, size TEXT,
        colour TEXT, qty TEXT, replacement INTEGER NOT NULL DEFAULT 0,
        created_at TEXT, created_by TEXT, UNIQUE (shipment_no, packing_no));
    CREATE TABLE IF NOT EXISTS warna_custom (nama TEXT PRIMARY KEY);
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL, salt TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'staff', created_at TEXT);
    CREATE TABLE IF NOT EXISTS reprint_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT, shipment_no TEXT, packing_no TEXT,
        alasan TEXT, dicetak_oleh TEXT, created_at TEXT);
    CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT, aksi TEXT, shipment_no TEXT,
        packing_no TEXT, detail TEXT, dilakukan_oleh TEXT, created_at TEXT);""")
    cols = [c[1] for c in conn.execute("PRAGMA table_info(packing)").fetchall()]
    if "replacement" not in cols:
        conn.execute("ALTER TABLE packing ADD COLUMN replacement INTEGER NOT NULL DEFAULT 0")
    if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        s, h = hash_password("admin123")
        conn.execute("INSERT INTO users (username,password_hash,salt,role,created_at) VALUES (?,?,?,?,?)",
                     ("admin", h, s, "admin", datetime.now().strftime("%d-%m-%Y %H:%M")))
    conn.commit(); return conn

def log_audit(conn, aksi, ship, pack, detail=""):
    conn.execute("INSERT INTO audit_log (aksi,shipment_no,packing_no,detail,dilakukan_oleh,created_at) "
                 "VALUES (?,?,?,?,?,?)", (aksi, ship, pack, detail, session.get("user", ""),
                 datetime.now().strftime("%d-%m-%Y %H:%M")))

def login_required(f):
    @wraps(f)
    def w(*a, **k):
        if "user" not in session: return jsonify(error="Belum login."), 401
        return f(*a, **k)
    return w

def kunci_urut(no):
    return [int(x) for x in re.split(r"[.\-]", str(no)) if x.isdigit()]

# ---------------- FILE STATIS (HTML/CSS/JS) ----------------
@app.route("/")
def f_index():
    return send_from_directory(BASE, "index.html")

@app.route("/style.css")
def f_css():
    return send_from_directory(BASE, "style.css")

@app.route("/app.js")
def f_js():
    return send_from_directory(BASE, "app.js")

# ---------------- AUTH ----------------
@app.route("/api/me")
def api_me():
    return jsonify(user=session.get("user"), role=session.get("role"))

@app.route("/api/login", methods=["POST"])
def api_login():
    d = request.json or {}
    conn = get_db()
    row = conn.execute("SELECT * FROM users WHERE username=?",
                       ((d.get("username") or "").strip(),)).fetchone()
    conn.close()
    if row and verifikasi_password(d.get("password") or "", row["salt"], row["password_hash"]):
        session["user"], session["role"] = row["username"], row["role"]
        return jsonify(ok=True, user=row["username"], role=row["role"])
    return jsonify(error="Nama pengguna atau kata sandi salah."), 401

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear(); return jsonify(ok=True)

@app.route("/api/ganti_sandi", methods=["POST"])
@login_required
def api_ganti():
    d = request.json or {}
    conn = get_db()
    me = conn.execute("SELECT * FROM users WHERE username=?", (session["user"],)).fetchone()
    if not verifikasi_password(d.get("lama", ""), me["salt"], me["password_hash"]):
        conn.close(); return jsonify(error="Kata sandi lama salah."), 400
    if len(d.get("baru", "")) < 6:
        conn.close(); return jsonify(error="Kata sandi baru minimal 6 karakter."), 400
    s, h = hash_password(d["baru"])
    conn.execute("UPDATE users SET password_hash=?,salt=? WHERE username=?", (h, s, session["user"]))
    conn.commit(); conn.close(); return jsonify(ok=True)

@app.route("/api/tambah_user", methods=["POST"])
@login_required
def api_tambah_user():
    if session.get("role") != "admin": return jsonify(error="Khusus admin."), 403
    d = request.json or {}
    if len(d.get("password", "")) < 6: return jsonify(error="Kata sandi minimal 6 karakter."), 400
    s, h = hash_password(d["password"])
    conn = get_db()
    try:
        conn.execute("INSERT INTO users (username,password_hash,salt,role,created_at) VALUES (?,?,?,?,?)",
                     ((d.get("username") or "").strip(), h, s, d.get("role", "staff"),
                      datetime.now().strftime("%d-%m-%Y %H:%M")))
        conn.commit(); conn.close(); return jsonify(ok=True)
    except sqlite3.IntegrityError:
        conn.close(); return jsonify(error="Username sudah dipakai."), 409

# ---------------- DATA REFERENSI ----------------
@app.route("/api/warna")
@login_required
def api_warna():
    conn = get_db()
    extra = [r["nama"] for r in conn.execute("SELECT nama FROM warna_custom ORDER BY nama").fetchall()]
    conn.close(); return jsonify(WARNA_OPSI + extra)

@app.route("/api/kode_buyer")
@login_required
def api_kode():
    conn = get_db()
    ships = [r["shipment_no"] for r in conn.execute("SELECT DISTINCT shipment_no FROM packing").fetchall()]
    conn.close()
    kode = set()
    for s in ships:
        m = re.match(r"[A-Za-z]+", s or "")
        kode.add(m.group(0) if m else (s or "").strip())
    return jsonify(sorted(k for k in kode if k))

@app.route("/api/next")
@login_required
def api_next():
    ship = request.args.get("shipment", "").strip()
    utama = request.args.get("utama", "").strip()
    jumlah = int(request.args.get("jumlah", "1"))
    if not ship or not utama.isdigit(): return jsonify(first=None)
    conn = get_db()
    cur = conn.execute("SELECT packing_no FROM packing WHERE shipment_no=? AND packing_no LIKE ?",
                       (ship, f"{utama}.%"))
    subs = []
    for r in cur.fetchall():
        try: subs.append(int(r["packing_no"].split(".")[1]))
        except (IndexError, ValueError): pass
    conn.close()
    start = (max(subs) + 1) if subs else 1
    return jsonify(first=f"{utama}.{start}", last=f"{utama}.{start + jumlah - 1}")

# ---------------- LABEL: LIST / BUAT / EDIT / HAPUS ----------------
@app.route("/api/labels")
@login_required
def api_labels():
    kode = request.args.get("kode", "")
    manual = request.args.get("manual", "")
    warna = request.args.get("warna", "")
    if manual:
        cond, params = "shipment_no LIKE ?", (manual + "%",)
    else:
        cond, params = "(shipment_no LIKE ? || '/%' OR shipment_no LIKE ? || ' %')", (kode, kode)
    conn = get_db()
    rows = conn.execute(f"SELECT * FROM packing WHERE {cond}", params).fetchall()
    conn.close()
    items = [dict(r) for r in rows]
    if warna and warna != "Semua warna":
        items = [i for i in items if (i.get("colour") or "").strip().lower() == warna.lower()]
    items.sort(key=lambda i: (i["shipment_no"], kunci_urut(i["packing_no"])))
    return jsonify(items)

@app.route("/api/labels", methods=["POST"])
@login_required
def api_create():
    d = request.json or {}
    ship = (d.get("shipment") or "").strip()
    utama = (d.get("utama") or "").strip()
    jumlah = int(d.get("jumlah") or 1)
    if not ship or not utama.isdigit():
        return jsonify(error="Shipment No dan Nomor Utama (angka) wajib diisi."), 400
    conn = get_db()
    if d.get("warna_baru"):
        conn.execute("INSERT OR IGNORE INTO warna_custom (nama) VALUES (?)", (d["warna_baru"],))
    cur = conn.execute("SELECT packing_no FROM packing WHERE shipment_no=? AND packing_no LIKE ?",
                       (ship, f"{utama}.%"))
    subs = []
    for r in cur.fetchall():
        try: subs.append(int(r["packing_no"].split(".")[1]))
        except (IndexError, ValueError): pass
    start = (max(subs) + 1) if subs else 1
    items = [{"packing_no": f"{utama}.{start+i}", "shipment_no": ship, "code": d.get("code", ""),
              "description": d.get("description", ""), "size": d.get("size", ""),
              "colour": d.get("colour", ""), "qty": d.get("qty", ""),
              "replacement": 1 if d.get("replacement") else 0} for i in range(jumlah)]
    try:
        conn.executemany("""INSERT INTO packing
            (packing_no,shipment_no,code,description,size,colour,qty,replacement,created_at,created_by)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            [(i["packing_no"], i["shipment_no"], i["code"], i["description"], i["size"], i["colour"],
              i["qty"], i["replacement"], datetime.now().strftime("%d-%m-%Y %H:%M"), session["user"])
             for i in items])
        conn.commit(); conn.close()
        return jsonify(ok=True, items=items)
    except sqlite3.IntegrityError:
        conn.rollback(); conn.close()
        return jsonify(error="Sebagian nomor sudah diterbitkan pengguna lain."), 409

@app.route("/api/labels/<int:id>", methods=["PUT"])
@login_required
def api_edit(id):
    d = request.json or {}
    conn = get_db()
    try:
        conn.execute("""UPDATE packing SET packing_no=?,shipment_no=?,code=?,description=?,
                        size=?,colour=?,qty=?,replacement=? WHERE id=?""",
                     (d["packing_no"], d["shipment_no"], d["code"], d["description"], d["size"],
                      d["colour"], d["qty"], 1 if d.get("replacement") else 0, id))
        log_audit(conn, "EDIT LABEL", d["shipment_no"], d["packing_no"], "via web")
        conn.commit(); conn.close(); return jsonify(ok=True)
    except sqlite3.IntegrityError:
        conn.rollback(); conn.close()
        return jsonify(error="Kombinasi Shipment + Packing No sudah dipakai."), 409

@app.route("/api/labels/<int:id>", methods=["DELETE"])
@login_required
def api_delete(id):
    conn = get_db()
    row = conn.execute("SELECT * FROM packing WHERE id=?", (id,)).fetchone()
    if row:
        conn.execute("DELETE FROM packing WHERE id=?", (id,))
        log_audit(conn, "HAPUS LABEL", row["shipment_no"], row["packing_no"],
                  f"description: {row['description']}")
        conn.commit()
    conn.close(); return jsonify(ok=True)

@app.route("/api/shipment", methods=["PUT"])
@login_required
def api_edit_ship():
    d = request.json or {}
    conn = get_db()
    try:
        conn.execute("UPDATE packing SET shipment_no=? WHERE shipment_no=?",
                     (d["baru"], d["lama"]))
        log_audit(conn, "EDIT SHIPMENT", d["baru"], "-", f"dari: {d['lama']}")
        conn.commit(); conn.close(); return jsonify(ok=True)
    except sqlite3.IntegrityError:
        conn.rollback(); conn.close()
        return jsonify(error="Sebagian nomor sudah ada di shipment tujuan."), 409

@app.route("/api/shipment", methods=["DELETE"])
@login_required
def api_delete_ship():
    ship = (request.json or {}).get("ship", "")
    conn = get_db()
    n = conn.execute("SELECT COUNT(*) c FROM packing WHERE shipment_no=?", (ship,)).fetchone()["c"]
    conn.execute("DELETE FROM packing WHERE shipment_no=?", (ship,))
    log_audit(conn, "HAPUS SHIPMENT", ship, "-", f"{n} koli dihapus")
    conn.commit(); conn.close(); return jsonify(ok=True, n=n)

# ---------------- CETAK ULANG & LOG ----------------
@app.route("/api/reprint", methods=["POST"])
@login_required
def api_reprint():
    d = request.json or {}
    conn = get_db()
    conn.execute("INSERT INTO reprint_log (shipment_no,packing_no,alasan,dicetak_oleh,created_at) "
                 "VALUES (?,?,?,?,?)", (d["shipment_no"], d["packing_no"], d.get("alasan", ""),
                 session["user"], datetime.now().strftime("%d-%m-%Y %H:%M")))
    conn.commit(); conn.close(); return jsonify(ok=True)

@app.route("/api/reprint_log")
@login_required
def api_reprint_log():
    conn = get_db()
    rows = conn.execute("SELECT * FROM reprint_log ORDER BY id DESC LIMIT 10").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

# ---------------- EXPORT EXCEL ----------------
def buat_excel_template(items):
    wb = Workbook(); ws = wb.active; ws.title = "Packing List"
    ws.column_dimensions["A"].width = 16; ws.column_dimensions["B"].width = 2
    ws.column_dimensions["C"].width = 45
    TEBAL = Side(style="medium", color="000000"); TIPIS = Side(style="thin", color="B0B0B0")
    prev_ship, count_in_ship = None, 0
    for idx, d in enumerate(items):
        r = 1 + idx * 10
        if d["shipment_no"] != prev_ship:
            if idx > 0: ws.row_breaks.append(Break(id=r - 1))
            prev_ship, count_in_ship = d["shipment_no"], 0
        elif count_in_ship % 2 == 0:
            ws.row_breaks.append(Break(id=r - 1))
        count_in_ship += 1
        for i in range(9): ws.row_dimensions[r + i].height = 22 if i < 3 else 24
        ws.row_dimensions[r + 9].height = 8
        for row in range(r, r + 9):
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = Border(
                    left=TEBAL if col == 1 else TIPIS, right=TEBAL if col == 3 else TIPIS,
                    top=TEBAL if row == r else TIPIS, bottom=TEBAL if row == r + 8 else TIPIS)
        ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=1)
        ws.merge_cells(start_row=r, start_column=2, end_row=r+2, end_column=2)
        ws.merge_cells(start_row=r, start_column=3, end_row=r+2, end_column=3)
        ws[f"A{r}"] = "PACKING NO"; ws[f"A{r}"].font = Font(name="Arial", size=13)
        ws[f"A{r}"].alignment = Alignment(vertical="center"); ws[f"B{r}"] = ":"
        ws[f"C{r}"] = d["packing_no"]; ws[f"C{r}"].font = Font(name="Arial", size=36, bold=True)
        ws[f"C{r}"].alignment = Alignment(horizontal="center", vertical="center")
        for i, (k, v) in enumerate([("SHIPMENT NO", d["shipment_no"]), ("CODE", d["code"]),
                                    ("DESCRIPTION", d["description"]), ("SIZE", d["size"]),
                                    ("COLOUR", d["colour"]), ("QTY", d["qty"])]):
            ws[f"A{r+3+i}"] = k; ws[f"A{r+3+i}"].font = Font(name="Arial", size=13)
            ws[f"A{r+3+i}"].alignment = Alignment(vertical="center"); ws[f"B{r+3+i}"] = ":"
            ws[f"C{r+3+i}"] = v; ws[f"C{r+3+i}"].font = Font(name="Arial", size=13, bold=True)
            ws[f"C{r+3+i}"].alignment = Alignment(vertical="center")
        if d.get("replacement"):
            ws[f"C{r+3}"] = CellRichText(
                TextBlock(InlineFont(rFont="Arial", sz=13, b=True), str(d["shipment_no"])),
                TextBlock(InlineFont(rFont="Arial", sz=13, b=True, color="FF0000"), "  # REPLACEMENT"))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

@app.route("/api/export", methods=["POST"])
@login_required
def api_export():
    items = (request.json or {}).get("items", [])
    buf = buat_excel_template(items)
    return send_file(buf, as_attachment=True, download_name="packing.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    print("=" * 60)
    print("  SISTEM MARKING CV JAVA VOLUME ART - http://localhost:5000")
    print("=" * 60)
    app.run(host="0.0.0.0", port=5000, debug=False)